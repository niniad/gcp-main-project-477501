import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = r'G:\マイドライブ\仕事\data\_old\MF総勘定元帳.xlsx'
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f'Sheets: {wb.sheetnames}')

for s in wb.sheetnames:
    ws = wb[s]
    # Check first row for headers
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        vals = [str(v)[:30] if v is not None else '' for v in row[:15]]
        print(f'  Sheet "{s}" row: {vals}')
    # Check if sheet name contains Amazon or 売掛
    if 'Amazon' in s or '売掛' in s or 'amazon' in s.lower():
        print(f'\n=== Found relevant sheet: {s} ===')
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True)):
            vals = [str(v)[:40] if v is not None else '' for v in row[:15]]
            print(f'  Row {i+1}: {vals}')
wb.close()
